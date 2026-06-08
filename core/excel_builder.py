"""
excel_builder.py
Builds a formatted Excel workbook from universal document extraction results.
Layout: one section header per document section, attributes listed below.
"""

from __future__ import annotations
from pathlib import Path
from typing import Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY    = "1B2A4A"
ORANGE  = "F47920"
WHITE   = "FFFFFF"
LIGHT   = "F5F7FA"
BORDER  = "DDE3EC"
MUTED   = "6C757D"

# Category → section header colour
CATEGORY_COLORS = {
    "identity":   ("1B2A4A", "DBEAFE"),  # navy header, blue section
    "academic":   ("166534", "DCFCE7"),  # green
    "employment": ("92400E", "FEF3C7"),  # amber
    "financial":  ("1E40AF", "EFF6FF"),  # blue
    "legal":      ("6B21A8", "F3E8FF"),  # purple
    "medical":    ("991B1B", "FEE2E2"),  # red
    "other":      ("374151", "F1F5F9"),  # grey
}

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _thin_border():
    s = Side(style="thin", color=BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _bottom_border():
    s = Side(style="thin", color=BORDER)
    return Border(bottom=s)


class ExcelBuilder:
    def build(self, results: list[dict[str, Any]], output_path: str | Path) -> Path:
        output_path = Path(output_path)
        wb = openpyxl.Workbook()

        # Summary sheet
        self._summary_sheet(wb, results)

        # Consolidated sheet (one row per document, all attributes as columns)
        self._consolidated_sheet(wb, results)

        # Per-document detail sheets
        for i, r in enumerate(results, start=1):
            name = r.get("_filename", f"Doc{i}")
            safe = name[:25].replace("/", "-").replace("\\", "-")
            self._detail_sheet(wb, r, sheet_name=f"{i}. {safe}")

        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

        wb.save(output_path)
        return output_path

    # ── Summary sheet ─────────────────────────────────────────────────────────

    def _summary_sheet(self, wb, results):
        ws = wb.active or wb.create_sheet("Summary")
        ws.title = "Summary"
        ws.sheet_view.showGridLines = False

        # Title
        ws.merge_cells("A1:F1")
        c = ws["A1"]
        c.value = "Document Intelligence — Extraction Summary"
        c.fill = _fill(NAVY)
        c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 34

        # Headers
        hdrs = ["#", "Filename", "Document Type", "Category", "Issuer", "Date"]
        for col, h in enumerate(hdrs, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.fill = _fill(ORANGE)
            c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin_border()
        ws.row_dimensions[2].height = 22

        for i, r in enumerate(results, 1):
            cat   = r.get("document_category", "other")
            _, bg = CATEGORY_COLORS.get(cat, ("374151", "F1F5F9"))
            row   = i + 2
            vals  = [
                i,
                r.get("_filename", ""),
                r.get("document_type", ""),
                cat.title(),
                r.get("issuer", ""),
                r.get("document_date", ""),
            ]
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = _fill(bg if col > 1 else WHITE)
                c.font = Font(size=10, name="Calibri", color=NAVY,
                              bold=(col == 1))
                c.alignment = Alignment(horizontal="left" if col > 1 else "center",
                                        vertical="center")
                c.border = _thin_border()
            ws.row_dimensions[row].height = 18

        for col, w in zip("ABCDEF", [5, 28, 28, 14, 28, 14]):
            ws.column_dimensions[col].width = w

    # ── Consolidated sheet ────────────────────────────────────────────────────

    def _consolidated_sheet(self, wb, results):
        """
        One row per document. Columns are dynamically built from every unique
        attribute name seen across all documents.
        Fixed leading columns: #, Filename, Document Type, Category, Issuer, Date.
        Then one column per unique attribute name (title-cased, de-duped).
        """
        ws = wb.create_sheet(title="Consolidated")
        ws.sheet_view.showGridLines = False

        # ── Pass 1: collect all unique attribute names in encounter order ─────
        seen_attrs: list[str] = []
        seen_set:   set[str]  = set()

        for r in results:
            for section in (r.get("sections") or []):
                for attr in (section.get("attributes") or []):
                    name = (attr.get("name") or "").strip()
                    key  = name.lower()
                    if name and key not in seen_set:
                        seen_attrs.append(name)
                        seen_set.add(key)

        FIXED_COLS = ["#", "Filename", "Document Type", "Category", "Issuer", "Document Date"]
        all_cols   = FIXED_COLS + seen_attrs
        n_fixed    = len(FIXED_COLS)

        # ── Title banner ──────────────────────────────────────────────────────
        last_col_letter = self._col_letter(len(all_cols))
        ws.merge_cells(f"A1:{last_col_letter}1")
        c = ws["A1"]
        c.value = "Consolidated View — All Documents"
        c.fill  = _fill(NAVY)
        c.font  = Font(bold=True, color=WHITE, size=13, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32

        # ── Header row ────────────────────────────────────────────────────────
        for col_idx, col_name in enumerate(all_cols, start=1):
            bg = ORANGE if col_idx <= n_fixed else "2D4A7A"
            c  = ws.cell(row=2, column=col_idx, value=col_name)
            c.fill      = _fill(bg)
            c.font      = Font(bold=True, color=WHITE, size=9, name="Calibri")
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border    = _thin_border()
        ws.row_dimensions[2].height = 28

        # ── Data rows ─────────────────────────────────────────────────────────
        for doc_idx, r in enumerate(results, start=1):
            row_num = doc_idx + 2
            cat     = r.get("document_category", "other")
            _, sec_bg = CATEGORY_COLORS.get(cat, ("374151", "F1F5F9"))

            # Build attribute lookup: lowercase name → value
            attr_lookup: dict[str, str] = {}
            for section in (r.get("sections") or []):
                for attr in (section.get("attributes") or []):
                    name = (attr.get("name") or "").strip()
                    val  = attr.get("value", "")
                    key  = name.lower()
                    if key and key not in attr_lookup:
                        attr_lookup[key] = str(val) if val is not None else ""

            # Fixed column values
            fixed_vals = [
                doc_idx,
                r.get("_filename", ""),
                r.get("document_type", ""),
                cat.title(),
                r.get("issuer", ""),
                r.get("document_date", ""),
            ]

            # Write fixed columns
            for col_idx, val in enumerate(fixed_vals, start=1):
                bg = LIGHT if doc_idx % 2 == 0 else WHITE
                c  = ws.cell(row=row_num, column=col_idx, value=val)
                c.fill      = _fill(bg)
                c.font      = Font(size=10, color=NAVY,
                                   bold=(col_idx == 1), name="Calibri")
                c.alignment = Alignment(horizontal="left" if col_idx > 1 else "center",
                                        vertical="center", wrap_text=True)
                c.border    = _thin_border()

            # Write dynamic attribute columns
            for col_idx, attr_name in enumerate(seen_attrs, start=n_fixed + 1):
                val = attr_lookup.get(attr_name.lower(), "")
                bg  = "EFF6FF" if doc_idx % 2 == 0 else WHITE
                c   = ws.cell(row=row_num, column=col_idx, value=val)
                c.fill      = _fill(bg)
                c.font      = Font(size=10, color="1F2937", name="Calibri")
                c.alignment = Alignment(horizontal="left", vertical="center",
                                        wrap_text=True)
                c.border    = _thin_border()

            ws.row_dimensions[row_num].height = 20

        # ── Column widths ─────────────────────────────────────────────────────
        fixed_widths = [4, 24, 26, 12, 24, 14]
        for i, w in enumerate(fixed_widths, start=1):
            ws.column_dimensions[self._col_letter(i)].width = w
        for i in range(n_fixed + 1, len(all_cols) + 1):
            ws.column_dimensions[self._col_letter(i)].width = 22

        # Freeze header rows and fixed columns
        ws.freeze_panes = "G3"

    @staticmethod
    def _col_letter(col_idx: int) -> str:
        """Convert 1-based column index to Excel letter(s): 1→A, 27→AA, etc."""
        result = ""
        while col_idx:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    # ── Detail sheet ──────────────────────────────────────────────────────────

    def _detail_sheet(self, wb, result: dict, sheet_name: str):
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False

        cat         = result.get("document_category", "other")
        hdr_color, sec_bg = CATEGORY_COLORS.get(cat, ("374151", "F1F5F9"))
        doc_type    = result.get("document_type", "Unknown Document")
        issuer      = result.get("issuer", "")
        doc_date    = result.get("document_date", "")
        filename    = result.get("_filename", "")
        sections    = result.get("sections", [])
        flags       = result.get("flags", {})

        # ── Document banner ──────────────────────────────────────────────────
        ws.merge_cells("A1:C1")
        c = ws["A1"]
        c.value = doc_type
        c.fill = _fill(hdr_color)
        c.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center",
                                indent=1)
        ws.row_dimensions[1].height = 36

        # Sub-banner: issuer | date | filename
        ws.merge_cells("A2:C2")
        meta_parts = []
        if issuer:   meta_parts.append(issuer)
        if doc_date: meta_parts.append(f"Date: {doc_date}")
        if filename: meta_parts.append(f"File: {filename}")
        c2 = ws["A2"]
        c2.value = "   ·   ".join(meta_parts)
        c2.fill = _fill("E8EDF4")
        c2.font = Font(size=9, color=NAVY, name="Calibri")
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[2].height = 18

        # ── Flags row ────────────────────────────────────────────────────────
        flag_items = [
            ("📷 Photo",     flags.get("has_photo")),
            ("✍️ Signature", flags.get("has_signature")),
            ("🔖 Stamp",     flags.get("has_stamp")),
            ("📲 QR Code",   flags.get("has_qr_code")),
            ("📊 Barcode",   flags.get("has_barcode")),
            ("🔒 Masked",    flags.get("is_masked")),
        ]
        flag_str = "   ".join(
            f"{label}: {'Yes' if val else 'No'}" for label, val in flag_items
        )
        ws.merge_cells("A3:C3")
        c3 = ws["A3"]
        c3.value = flag_str
        c3.fill = _fill(WHITE)
        c3.font = Font(size=9, color=MUTED, name="Calibri")
        c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[3].height = 16

        current_row = 5  # leave row 4 as spacer

        # ── Sections ─────────────────────────────────────────────────────────
        if result.get("_parse_error") and not sections:
            ws.merge_cells(f"A{current_row}:C{current_row}")
            ec = ws[f"A{current_row}"]
            ec.value = f"⚠ Extraction error: {result.get('_error_message', 'Unknown error')}"
            ec.fill = _fill("FEE2E2")
            ec.font = Font(color="991B1B", size=10, name="Calibri")
            ec.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        else:
            for section in sections:
                section_title = section.get("section_title", "Details")
                attributes    = section.get("attributes", [])

                # Section header row
                ws.merge_cells(f"A{current_row}:C{current_row}")
                sh = ws[f"A{current_row}"]
                sh.value = section_title.upper()
                sh.fill = _fill(sec_bg)
                sh.font = Font(bold=True, size=10, color=hdr_color, name="Calibri")
                sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                ws.row_dimensions[current_row].height = 22
                current_row += 1

                # Column sub-headers (Attribute | Value | Confidence)
                for col, (h, w) in enumerate(
                    [("Attribute", 28), ("Value", 40), ("Confidence", 12)], 1
                ):
                    c = ws.cell(row=current_row, column=col, value=h)
                    c.fill = _fill(NAVY)
                    c.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
                    c.alignment = Alignment(horizontal="left" if col < 3 else "center",
                                            vertical="center", indent=1)
                    c.border = _thin_border()
                ws.row_dimensions[current_row].height = 18
                current_row += 1

                # Attribute rows
                for j, attr in enumerate(attributes):
                    bg = "FFFFFF" if j % 2 == 0 else LIGHT
                    name_val = attr.get("name", "")
                    val      = attr.get("value", "")
                    conf     = (attr.get("confidence") or "high").lower()

                    conf_color = {
                        "high":   ("DCFCE7", "166534"),
                        "medium": ("FEF3C7", "92400E"),
                        "low":    ("FEE2E2", "991B1B"),
                    }.get(conf, ("F1F5F9", "374151"))

                    # Attribute name
                    nc = ws.cell(row=current_row, column=1, value=name_val)
                    nc.fill = _fill(bg)
                    nc.font = Font(bold=True, size=10, color=NAVY, name="Calibri")
                    nc.alignment = Alignment(horizontal="left", vertical="center",
                                             wrap_text=True, indent=1)
                    nc.border = _thin_border()

                    # Value
                    vc = ws.cell(row=current_row, column=2, value=str(val) if val is not None else "")
                    vc.fill = _fill(bg)
                    vc.font = Font(size=10, color="1F2937", name="Calibri")
                    vc.alignment = Alignment(horizontal="left", vertical="center",
                                             wrap_text=True, indent=1)
                    vc.border = _thin_border()

                    # Confidence
                    cbg, cfg = conf_color
                    cc = ws.cell(row=current_row, column=3, value=conf.upper())
                    cc.fill = _fill(cbg)
                    cc.font = Font(bold=True, size=9, color=cfg, name="Calibri")
                    cc.alignment = Alignment(horizontal="center", vertical="center")
                    cc.border = _thin_border()

                    ws.row_dimensions[current_row].height = 18
                    current_row += 1

                # Spacer row between sections
                current_row += 1

        # ── Raw text at the bottom ────────────────────────────────────────────
        raw = result.get("raw_text", "")
        if raw:
            current_row += 1
            ws.merge_cells(f"A{current_row}:C{current_row}")
            rh = ws[f"A{current_row}"]
            rh.value = "RAW EXTRACTED TEXT"
            rh.fill = _fill(NAVY)
            rh.font = Font(bold=True, color=WHITE, size=9, name="Calibri")
            rh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[current_row].height = 16
            current_row += 1

            ws.merge_cells(f"A{current_row}:C{current_row + 5}")
            rt = ws[f"A{current_row}"]
            rt.value = raw
            rt.font = Font(size=9, name="Courier New", color="374151")
            rt.fill = _fill("F8FAFC")
            rt.alignment = Alignment(horizontal="left", vertical="top",
                                     wrap_text=True)
            ws.row_dimensions[current_row].height = 100

        # Column widths
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 13
